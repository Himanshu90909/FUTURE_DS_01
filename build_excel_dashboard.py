from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import pandas as pd

ROOT = Path(__file__).resolve().parent
DATA_PATH = ROOT / 'superstore_sales.csv'
OUTPUT_PATH = ROOT / 'FUTURE_DS_01_Excel_Dashboard.xlsx'
NAVY, BLUE, TEAL, AMBER, WHITE, SLATE = '17324D', '2563EB', '0F766E', 'F59E0B', 'FFFFFF', '475569'
REQUIRED_COLUMNS = {
    'Order ID', 'Order Date', 'Ship Date', 'Customer ID', 'Product ID',
    'Category', 'Sub-Category', 'Region', 'Segment', 'Ship Mode',
    'Sales', 'Quantity', 'Discount', 'Profit',
}


def style_header(cells, fill=BLUE):
    for cell in cells:
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(name='Aptos', bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def add_table(ws, ref, name, style='TableStyleMedium2'):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def title(ws, ref, text):
    ws.merge_cells(ref)
    c = ws[ref.split(':')[0]]
    c.value = text
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.font = Font(name='Aptos Display', size=18, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')


def load_data():
    """Load and normalize the transaction-level source data."""
    df = pd.read_csv(DATA_PATH)
    df.columns = df.columns.str.strip()

    missing = REQUIRED_COLUMNS.difference(df.columns)
    if missing:
        raise ValueError(
            f'Missing required columns in {DATA_PATH.name}: '
            f'{", ".join(sorted(missing))}'
        )

    for column in ['Order Date', 'Ship Date']:
        df[column] = pd.to_datetime(df[column], errors='coerce')
    for column in ['Sales', 'Quantity', 'Discount', 'Profit']:
        df[column] = pd.to_numeric(df[column], errors='coerce')

    # Store the margin as a decimal so Excel's 0.0% format renders correctly.
    df['Profit Margin %'] = (
        df['Profit'].div(df['Sales'].where(df['Sales'].ne(0))).fillna(0)
    )
    return df


def main():
    df = load_data()
    df = df.dropna(subset=['Order Date', 'Sales', 'Profit', 'Quantity']).copy()
    df['Month'] = df['Order Date'].dt.to_period('M').astype(str)

    region = df.groupby('Region', as_index=False).agg(Orders=('Order ID','nunique'), Sales=('Sales','sum'), Profit=('Profit','sum'), Quantity=('Quantity','sum'), Avg_Discount=('Discount','mean')).sort_values('Sales', ascending=False)
    subcat = df.groupby(['Category','Sub-Category'], as_index=False).agg(Orders=('Order ID','nunique'), Sales=('Sales','sum'), Profit=('Profit','sum'), Quantity=('Quantity','sum'), Avg_Discount=('Discount','mean')).sort_values('Profit', ascending=False)
    product = df.groupby(['Product ID','Product Name','Category','Sub-Category'], as_index=False).agg(Orders=('Order ID','nunique'), Sales=('Sales','sum'), Profit=('Profit','sum'), Quantity=('Quantity','sum')).sort_values('Sales', ascending=False)
    monthly = df.groupby('Month', as_index=False).agg(Sales=('Sales','sum'), Profit=('Profit','sum'), Orders=('Order ID','nunique'), Quantity=('Quantity','sum'))
    segment = df.groupby('Segment', as_index=False).agg(Orders=('Order ID','nunique'), Sales=('Sales','sum'), Profit=('Profit','sum'), Quantity=('Quantity','sum'))
    ship = df.groupby('Ship Mode', as_index=False).agg(Orders=('Order ID','nunique'), Sales=('Sales','sum'), Profit=('Profit','sum'))

    wb = Workbook(); dashboard = wb.active; dashboard.title = 'Dashboard'
    clean = wb.create_sheet('Clean_Data'); region_ws = wb.create_sheet('Region_Summary'); subcat_ws = wb.create_sheet('SubCategory_Summary'); product_ws = wb.create_sheet('Top_Products'); monthly_ws = wb.create_sheet('Monthly_Trend'); method = wb.create_sheet('PowerBI_Guide')

    # Clean data
    clean.append(list(df.columns))
    for r in df.itertuples(index=False, name=None): clean.append(list(r))
    style_header(clean[1], TEAL); clean.freeze_panes = 'A2'; clean.auto_filter.ref = clean.dimensions
    add_table(clean, f'A1:{get_column_letter(clean.max_column)}{clean.max_row}', 'SalesData', 'TableStyleMedium4')
    widths = {'A':16,'B':12,'C':12,'D':16,'E':14,'F':22,'G':13,'H':18,'I':18,'J':16,'K':12,'L':10,'M':16,'N':16,'O':23,'P':64,'Q':14,'R':10,'S':12,'T':14,'U':14,'V':16,'W':12}
    for col, width in widths.items(): clean.column_dimensions[col].width = width
    for row in range(2, clean.max_row + 1):
        clean.cell(row, 2).number_format = 'yyyy-mm-dd'; clean.cell(row, 3).number_format = 'yyyy-mm-dd'; clean.cell(row, 17).number_format = '$#,##0.00'; clean.cell(row, 19).number_format = '0.0%'; clean.cell(row, 20).number_format = '$#,##0.00'; clean.cell(row, 21).number_format = '0.0%'

    def summary_sheet(ws, data, headers, table_name, fill):
        ws.append(headers)
        for r in data.itertuples(index=False, name=None): ws.append(list(r))
        style_header(ws[1], fill); add_table(ws, f'A1:{get_column_letter(ws.max_column)}{ws.max_row}', table_name, 'TableStyleMedium2'); ws.freeze_panes = 'A2'
        for col in range(1, ws.max_column + 1): ws.column_dimensions[get_column_letter(col)].width = 20
        header_columns = {cell.value: cell.column for cell in ws[1]}
        for name in ['Sales', 'Profit']:
            if name in header_columns:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, header_columns[name]).number_format = '$#,##0.00'
        for name in ['Avg_Discount', 'Profit Margin %']:
            if name in header_columns:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, header_columns[name]).number_format = '0.0%'
        for name in ['Orders', 'Quantity']:
            if name in header_columns:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, header_columns[name]).number_format = '#,##0'
        return ws

    summary_sheet(region_ws, region, list(region.columns), 'RegionSummary', BLUE)
    summary_sheet(subcat_ws, subcat, list(subcat.columns), 'SubCategorySummary', TEAL)
    summary_sheet(product_ws, product.head(25), list(product.columns), 'TopProducts', AMBER)
    summary_sheet(monthly_ws, monthly, list(monthly.columns), 'MonthlyTrend', BLUE)
    subcat_ws.conditional_formatting.add(f'D2:D{subcat_ws.max_row}', ColorScaleRule(start_type='min', start_color='FECACA', mid_type='percentile', mid_value=50, mid_color='FEF3C7', end_type='max', end_color='BBF7D0'))

    # Power BI guide sheet
    title(method, 'A1:H2', 'FUTURE_DS_01 | Power BI Model & Measures')
    guide = [
        ('Source table', 'Load superstore_sales.csv into Power BI as the SalesData table.'),
        ('Recommended visuals', 'Cards: Sales, Profit, Profit Margin, Orders. Line chart: Sales by Month. Filled map or bar chart: Sales and Profit by Region. Clustered bar: Sales vs Profit by Category/Sub-Category. Matrix: Product Name, Sales, Profit, Quantity.'),
        ('Relationships', 'For this deliverable, SalesData is a single flat fact table. Add a Calendar table related to SalesData[Order Date] when using time-intelligence measures.'),
        ('Refresh', 'The CSV is Power BI-ready: dates are parseable, numeric measures are separated from descriptive dimensions, and the source contains region, category, sub-category, product, segment, ship mode, sales, discount, quantity, and profit.'),
        ('Limitation', 'The source is a sample Superstore dataset. Treat findings as portfolio analysis, not audited company financials.'),
    ]
    method['A4'] = 'Topic'; method['B4'] = 'Guidance'; style_header(method[4], TEAL)
    for row in guide: method.append(list(row))
    for r in range(5, 5 + len(guide)): method[f'A{r}'].font = Font(bold=True, color=NAVY); method[f'B{r}'].alignment = Alignment(wrap_text=True, vertical='top'); method.row_dimensions[r].height = 42
    method.column_dimensions['A'].width = 22; method.column_dimensions['B'].width = 110

    # Dashboard
    dashboard.sheet_view.showGridLines = False; title(dashboard, 'A1:N2', 'FUTURE_DS_01 | Business Sales Performance Dashboard')
    dashboard.merge_cells('A3:N3'); dashboard['A3'] = f'Superstore transactions | {df["Order ID"].nunique():,} orders | {df["Order Date"].min().date()} to {df["Order Date"].max().date()} | Power BI-ready source'; dashboard['A3'].font = Font(italic=True, color=SLATE)
    kpis = [('B5:D5','B6:D7','Sales',df['Sales'].sum(),'E8F0FE',BLUE,'$#,##0'),('E5:G5','E6:G7','Profit',df['Profit'].sum(),'E6F5F1',TEAL,'$#,##0'),('H5:J5','H6:J7','Profit Margin',df['Profit'].sum()/df['Sales'].sum(),'FEF3C7',AMBER,'0.0%'),('K5:M5','K6:M7','Orders',df['Order ID'].nunique(),'E8F0FE',BLUE,'#,##0')]
    for lr, vr, label, value, fill, accent, fmt in kpis:
        dashboard.merge_cells(lr); dashboard.merge_cells(vr); lc=dashboard[lr.split(':')[0]]; vc=dashboard[vr.split(':')[0]]; lc.value=label; vc.value=value; lc.fill=PatternFill('solid',fgColor=fill); vc.fill=PatternFill('solid',fgColor=fill); lc.font=Font(bold=True,color=accent); vc.font=Font(bold=True,color=NAVY,size=20); lc.alignment=Alignment(horizontal='center'); vc.alignment=Alignment(horizontal='center',vertical='center'); vc.number_format=fmt
    dashboard.merge_cells('B9:G9'); dashboard['B9']='Regional sales and profit'; dashboard['B9'].font=Font(bold=True,color=NAVY,size=12)
    dashboard.merge_cells('I9:N9'); dashboard['I9']='Monthly sales trend'; dashboard['I9'].font=Font(bold=True,color=NAVY,size=12)
    dashboard['P1']='Region'; dashboard['Q1']='Sales'; dashboard['R1']='Profit';
    for i, r in enumerate(region.itertuples(index=False),2): dashboard.cell(i,16).value=r.Region; dashboard.cell(i,17).value=r.Sales; dashboard.cell(i,18).value=r.Profit
    dashboard['T1']='Month'; dashboard['U1']='Sales';
    for i, r in enumerate(monthly.itertuples(index=False),2): dashboard.cell(i,20).value=r.Month; dashboard.cell(i,21).value=r.Sales
    for c in range(16,22): dashboard.column_dimensions[get_column_letter(c)].hidden=True
    chart1=BarChart(); chart1.type='col'; chart1.style=10; chart1.title='Sales by Region'; chart1.y_axis.title='Sales ($)'; chart1.height=8; chart1.width=12; chart1.add_data(Reference(dashboard,min_col=17,max_col=18,min_row=1,max_row=1+len(region)),titles_from_data=True); chart1.set_categories(Reference(dashboard,min_col=16,min_row=2,max_row=1+len(region))); dashboard.add_chart(chart1,'B10')
    chart2=LineChart(); chart2.style=13; chart2.title='Monthly Sales'; chart2.y_axis.title='Sales ($)'; chart2.height=8; chart2.width=12; chart2.add_data(Reference(dashboard,min_col=21,min_row=1,max_row=1+len(monthly)),titles_from_data=True); chart2.set_categories(Reference(dashboard,min_col=20,min_row=2,max_row=1+len(monthly))); dashboard.add_chart(chart2,'I10')
    dashboard.merge_cells('B28:N28'); dashboard['B28']='Executive recommendations'; dashboard['B28'].fill=PatternFill('solid',fgColor=NAVY); dashboard['B28'].font=Font(bold=True,color=WHITE,size=12)
    recommendations=[f'1. Focus regional growth and inventory planning on {region.iloc[0]["Region"]}, the leading sales region.',f'2. Review {subcat.sort_values("Profit").iloc[0]["Sub-Category"]} for pricing, discount, and fulfillment-cost issues because it has the lowest profit.',f'3. Use the Power BI model to filter Sales, Profit, Quantity, and Margin by date, region, category, sub-category, segment, ship mode, and product.', '4. Add a Calendar table in Power BI for year-over-year growth, month-over-month change, and rolling profit measures.']
    for i, text in enumerate(recommendations,29): dashboard.merge_cells(start_row=i,start_column=2,end_row=i,end_column=14); dashboard.cell(i,2).value=text; dashboard.cell(i,2).alignment=Alignment(wrap_text=True); dashboard.cell(i,2).font=Font(color=SLATE); dashboard.row_dimensions[i].height=28
    dashboard.merge_cells('B35:N35'); dashboard['B35']='Dataset note: this is a sample Superstore dataset for portfolio and dashboard practice, not audited company financial data.'; dashboard['B35'].font=Font(italic=True,color=SLATE,size=9)
    for col in range(1,15): dashboard.column_dimensions[get_column_letter(col)].width = 15
    dashboard.column_dimensions['A'].width=3; dashboard.column_dimensions['H'].width=4; dashboard.freeze_panes='B5'; dashboard.print_area='A1:N35'; dashboard.page_setup.fitToWidth=1; dashboard.page_setup.fitToHeight=1
    wb.active=0; wb.save(OUTPUT_PATH); print(f'Created {OUTPUT_PATH}')


if __name__ == '__main__': main()
